import openpyxl
from datetime import datetime
import formatting
import dbx_aws


def age_format(date):
    date_diff = datetime.now() - date
    days = date_diff.days
    return days

def row_format(row):
    new_row = []
    new_row.append(row[0]) # barrel_no
    new_row.append(row[1].title()) # type
    # gallons
    if row[5] == 'Tasters Club':
        new_row.append(59)
    elif row[2] > 40:
        new_row.append(53)
    elif row[2] > 20 and row[2] < 40:
        new_row.append(30)
    else:
        new_row.append(15)
    new_row.append(row[2]) # pg
    new_row.append(row[3]) # date
    new_row.append(age_format(row[3])) # age
    if row[5] is None:
        new_row.append('')
    else:
        new_row.append(row[5]) # investor
    return new_row



wb = openpyxl.load_workbook('barrels.xlsx')
sheets = wb.sheetnames

# Whiskey transfer
ws = wb['78 Montgomery Rum']
barrels = []
for row in ws.iter_rows(min_row=2, values_only=True):
    barrels.append(row_format(row))

conn = dbx_aws.db_conn()
cur = conn.cursor()
query = "INSERT INTO barrels VALUES (%s,%s,%s,%s,%s,%s,%s)"
for row in barrels:
    cur.execute(query, row)
conn.commit()
conn.close()
